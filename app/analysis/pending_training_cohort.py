from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import date, timedelta
from pathlib import Path

from app.analysis.strategy import Strategy
from app.marketdata.price_history_provider import PriceHistoryProvider


@dataclass(frozen=True, slots=True)
class PendingTradeEntry:
    symbol: str
    decision_date: date
    expiration: date
    reference_price: float
    strategy: Strategy
    score: int | None
    short_put_strike: float
    short_call_strike: float
    long_put_strike: float | None = None
    long_call_strike: float | None = None

    def __post_init__(self) -> None:
        symbol = self.symbol.strip().upper()
        if not symbol:
            raise ValueError("symbol must not be empty")
        if self.reference_price <= 0:
            raise ValueError("reference_price must be greater than zero")
        if self.short_put_strike >= self.short_call_strike:
            raise ValueError("short put strike must be below short call strike")
        object.__setattr__(self, "symbol", symbol)


@dataclass(frozen=True, slots=True)
class FirstReactionSnapshot:
    symbol: str
    decision_date: date
    expiration: date
    reference_price: float
    strategy: Strategy
    score: int | None
    short_put_strike: float
    short_call_strike: float
    long_put_strike: float | None
    long_call_strike: float | None
    reaction_date: date
    open: float
    high: float
    low: float
    close: float
    volume: int
    gap_percent: float
    close_move_percent: float
    high_move_percent: float
    low_move_percent: float
    put_touched: bool
    call_touched: bool
    finished_inside_short_strikes: bool


@dataclass(frozen=True, slots=True)
class PendingTrainingCohort:
    decision_date: date
    source_file: str
    captured_trades: tuple[FirstReactionSnapshot, ...]
    missing_symbols: tuple[str, ...] = ()
    schema_version: int = 1


class PendingTradeExcelImporter:
    REQUIRED_HEADERS = {
        "Aktie",
        "Kurs",
        "Strategie",
        "ShortPutStrike",
        "ShortCallStrike",
    }

    def load(
        self,
        path: str | Path,
        decision_date: date,
    ) -> tuple[PendingTradeEntry, ...]:
        from openpyxl import load_workbook

        source = Path(path)
        workbook = load_workbook(source, read_only=True, data_only=True)
        worksheet = workbook["Earnings Crush"]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows)
        indexes = {
            str(value): index
            for index, value in enumerate(headers)
            if value is not None
        }
        missing = self.REQUIRED_HEADERS.difference(indexes)
        if missing:
            raise ValueError(
                "Missing required Excel columns: " + ", ".join(sorted(missing))
            )

        expiration = self._friday_of_week(decision_date)
        entries: list[PendingTradeEntry] = []
        for row in rows:
            symbol = row[indexes["Aktie"]]
            if symbol is None or not str(symbol).strip():
                continue

            entries.append(
                PendingTradeEntry(
                    symbol=str(symbol),
                    decision_date=decision_date,
                    expiration=expiration,
                    reference_price=float(row[indexes["Kurs"]]),
                    strategy=Strategy(str(row[indexes["Strategie"]])),
                    score=self._optional_int(row, indexes, "Score"),
                    short_put_strike=float(row[indexes["ShortPutStrike"]]),
                    short_call_strike=float(row[indexes["ShortCallStrike"]]),
                    long_put_strike=self._optional_float(
                        row, indexes, "LongPutStrike"
                    ),
                    long_call_strike=self._optional_float(
                        row, indexes, "LongCallStrike"
                    ),
                )
            )
        return tuple(entries)

    @staticmethod
    def _optional_float(row, indexes, name: str) -> float | None:
        index = indexes.get(name)
        if index is None or row[index] is None:
            return None
        return float(row[index])

    @staticmethod
    def _optional_int(row, indexes, name: str) -> int | None:
        index = indexes.get(name)
        if index is None or row[index] is None:
            return None
        return int(row[index])

    @staticmethod
    def _friday_of_week(day: date) -> date:
        return day + timedelta(days=(4 - day.weekday()) % 7)


class PendingTrainingCohortBuilder:
    """Capture the first completed trading day after the entry decision.

    This works for both earnings reported after the decision-day close and
    earnings reported before the next session opens. It deliberately ignores
    the wall-clock time at which the command is run.
    """

    def __init__(
        self,
        price_history_provider: PriceHistoryProvider,
        lookahead_calendar_days: int = 7,
    ) -> None:
        if lookahead_calendar_days <= 0:
            raise ValueError("lookahead_calendar_days must be greater than zero")
        self.price_history_provider = price_history_provider
        self.lookahead_calendar_days = lookahead_calendar_days

    def build(
        self,
        entries: tuple[PendingTradeEntry, ...],
        source_file: str,
    ) -> PendingTrainingCohort:
        if not entries:
            raise ValueError("entries must not be empty")
        decision_dates = {entry.decision_date for entry in entries}
        if len(decision_dates) != 1:
            raise ValueError("all entries must use the same decision date")

        captured: list[FirstReactionSnapshot] = []
        missing: list[str] = []
        for entry in entries:
            bars = self.price_history_provider.get_daily_bars(
                entry.symbol,
                entry.decision_date + timedelta(days=1),
                entry.decision_date
                + timedelta(days=self.lookahead_calendar_days),
            )
            first_bar = next(
                (bar for bar in bars if bar.date > entry.decision_date),
                None,
            )
            if first_bar is None:
                missing.append(entry.symbol)
                continue
            captured.append(self._snapshot(entry, first_bar))

        return PendingTrainingCohort(
            decision_date=entries[0].decision_date,
            source_file=source_file,
            captured_trades=tuple(captured),
            missing_symbols=tuple(missing),
        )

    @staticmethod
    def _snapshot(entry, bar) -> FirstReactionSnapshot:
        reference = entry.reference_price
        return FirstReactionSnapshot(
            symbol=entry.symbol,
            decision_date=entry.decision_date,
            expiration=entry.expiration,
            reference_price=reference,
            strategy=entry.strategy,
            score=entry.score,
            short_put_strike=entry.short_put_strike,
            short_call_strike=entry.short_call_strike,
            long_put_strike=entry.long_put_strike,
            long_call_strike=entry.long_call_strike,
            reaction_date=bar.date,
            open=bar.open,
            high=bar.high,
            low=bar.low,
            close=bar.close,
            volume=bar.volume,
            gap_percent=(bar.open / reference - 1) * 100,
            close_move_percent=(bar.close / reference - 1) * 100,
            high_move_percent=(bar.high / reference - 1) * 100,
            low_move_percent=(bar.low / reference - 1) * 100,
            put_touched=bar.low <= entry.short_put_strike,
            call_touched=bar.high >= entry.short_call_strike,
            finished_inside_short_strikes=(
                entry.short_put_strike < bar.close < entry.short_call_strike
            ),
        )


class PendingTrainingCohortRepository:
    def save(
        self,
        cohort: PendingTrainingCohort,
        path: str | Path,
    ) -> None:
        destination = Path(path)
        destination.parent.mkdir(parents=True, exist_ok=True)
        payload = asdict(cohort)
        payload["decision_date"] = cohort.decision_date.isoformat()
        for row, snapshot in zip(
            payload["captured_trades"], cohort.captured_trades
        ):
            row["decision_date"] = snapshot.decision_date.isoformat()
            row["expiration"] = snapshot.expiration.isoformat()
            row["reaction_date"] = snapshot.reaction_date.isoformat()
            row["strategy"] = snapshot.strategy.value
        destination.write_text(
            json.dumps(payload, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
