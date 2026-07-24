from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
import re

from openpyxl import load_workbook

from app.analysis.daily_trade_window_selector import DailyTradeWindowSelector
from app.analysis.strategy import Strategy
from app.analysis.trade_snapshot_review import TradeSnapshot
from app.marketdata.savvytrader_earnings_calendar_provider import (
    SavvyTraderEarningsCalendarProvider,
)


class TradeExcelSnapshotLoader:
    """Convert daily trade exports into reviewable snapshots."""

    DATE_PATTERN = re.compile(r"(\d{4}-\d{2}-\d{2})")

    def __init__(self, calendar_provider=None) -> None:
        self.calendar_provider = (
            calendar_provider or SavvyTraderEarningsCalendarProvider()
        )
        self.window_selector = DailyTradeWindowSelector()

    def load(self, paths: tuple[str | Path, ...]) -> tuple[TradeSnapshot, ...]:
        snapshots: list[TradeSnapshot] = []
        for path in paths:
            source = Path(path)
            trade_date = self._trade_date(source)
            events = self._events_for_trade_date(trade_date)
            event_by_symbol = {event.symbol.upper(): event for event in events}

            workbook = load_workbook(source, read_only=True, data_only=True)
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                continue
            indexes = {str(value): index for index, value in enumerate(headers)}

            for row in rows:
                symbol = str(row[indexes["Aktie"]] or "").strip().upper()
                if not symbol:
                    continue
                event = event_by_symbol.get(symbol)
                if event is None:
                    raise ValueError(
                        f"No matching earnings event for {symbol} on {trade_date}"
                    )
                report_date = event.report_date
                snapshots.append(
                    TradeSnapshot(
                        symbol=symbol,
                        decision_date=trade_date,
                        report_date=report_date,
                        expiration=self._friday_of_week(report_date),
                        strategy=Strategy(str(row[indexes["Strategie"]])),
                        reference_price=float(row[indexes["Kurs"]]),
                        short_put_strike=float(row[indexes["ShortPutStrike"]]),
                        short_call_strike=float(row[indexes["ShortCallStrike"]]),
                    )
                )
        return tuple(snapshots)

    def _events_for_trade_date(self, trade_date: date):
        next_date = self.window_selector.next_trading_weekday(trade_date)
        events = self.calendar_provider.get_events(
            trade_date,
            next_date + timedelta(days=1),
        )
        return self.window_selector.select(events, trade_date).events

    @classmethod
    def _trade_date(cls, path: Path) -> date:
        match = cls.DATE_PATTERN.search(path.name)
        if match is None:
            raise ValueError(f"Could not determine trade date from {path.name}")
        return date.fromisoformat(match.group(1))

    @staticmethod
    def _friday_of_week(value: date) -> date:
        return value + timedelta(days=4 - value.weekday())
