from datetime import date

from openpyxl import load_workbook

from app.analysis.earnings_crush_candidate import EarningsCrushCandidate
from app.analysis.strike_selection import StrikeSelection
from app.marketdata.models import EarningsEvent, MarketSnapshot, OptionQuote, Quote
from app.run_daily_trades import run_daily


class _CalendarProvider:
    def get_events(self, start_date, end_date):
        return [
            EarningsEvent(
                symbol="AAA",
                report_date=start_date,
                timing="AMC",
                source="test",
            ),
            EarningsEvent(
                symbol="BBB",
                report_date=date(2026, 7, 21),
                timing="BMO",
                source="test",
            ),
            EarningsEvent(
                symbol="IGNORED",
                report_date=start_date,
                timing="BMO",
                source="test",
            ),
        ]


class _Analyzer:
    def create_candidates(self, events):
        return [_candidate(events[0])]


class _AnalyzerFactory:
    def create(self, market_data):
        return _Analyzer()


def _candidate(event: EarningsEvent) -> EarningsCrushCandidate:
    candidate = EarningsCrushCandidate(
        earnings_event=event,
        snapshot=MarketSnapshot(
            symbol=event.symbol,
            quote=Quote(event.symbol, 100.0, "USD", "test"),
            news=[],
        ),
    )
    candidate.strike_selection = StrikeSelection(
        put=OptionQuote(
            symbol=event.symbol,
            expiration=date(2026, 7, 24),
            strike=90.0,
            option_type="put",
        ),
        call=OptionQuote(
            symbol=event.symbol,
            expiration=date(2026, 7, 24),
            strike=110.0,
            option_type="call",
        ),
        put_target=90.0,
        call_target=110.0,
    )
    return candidate


def test_run_daily_executes_full_pipeline_and_writes_excel(tmp_path):
    output_path = tmp_path / "daily_trades.xlsx"

    candidates = run_daily(
        trade_date=date(2026, 7, 20),
        output_path=output_path,
        calendar_provider=_CalendarProvider(),
        market_data=object(),
        analyzer_factory=_AnalyzerFactory(),
    )

    assert [candidate.earnings_event.symbol for candidate in candidates] == [
        "AAA",
        "BBB",
    ]
    assert output_path.exists()

    workbook = load_workbook(output_path, read_only=True)
    worksheet = workbook["Earnings Crush"]
    rows = list(worksheet.iter_rows(values_only=True))

    assert rows[0][:4] == ("Aktie", "Kurs", "Strategie", "Score")
    assert rows[1][0] == "AAA"
    assert rows[2][0] == "BBB"
    assert rows[1][2] == "Short Strangle"
    assert rows[2][2] == "Short Strangle"
